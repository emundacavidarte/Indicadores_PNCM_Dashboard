import openpyxl
import json

excel_path = "INDICADORES HIS - NIÑOS v2.1 - Junio 2026.xlsx"
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=False)

print("=== SHEET 1: INDICADORES ===")
ws = wb["INDICADORES"]
for r in range(1, 54):
    row_vals = [ws.cell(r, c).value for c in range(1, 34)]
    if any(row_vals):
        non_empty = [(c, str(val)[:50]) for c, val in enumerate(row_vals, 1) if val is not None]
        print(f"Row {r:2d}: {non_empty}")

print("\n=== SHEET 3: INDICADORES X DISTRITO ===")
ws_dist = wb["INDICADORES X DISTRITO"]
for r in range(1, 39):
    row_vals = [ws_dist.cell(r, c).value for c in range(1, 53)]
    if any(row_vals):
        non_empty = [(c, str(val)[:50]) for c, val in enumerate(row_vals, 1) if val is not None]
        print(f"Row {r:2d}: {non_empty[:10]}")

print("\n=== SHEET 6: Tabla - Columns Header ===")
ws_tabla = wb["Tabla"]
header = [ws_tabla.cell(1, c).value for c in range(1, 42)]
print(f"Tabla total columns: {len(header)}")
for idx, col in enumerate(header, 1):
    print(f"  Col {idx:2d}: {col}")

print("\n=== SHEET 7: config ===")
ws_cfg = wb["config"]
for r in range(1, 44):
    row_vals = [ws_cfg.cell(r, c).value for c in range(1, 5)]
    if any(row_vals):
        print(f"Row {r:2d}: {row_vals}")

wb.close()
