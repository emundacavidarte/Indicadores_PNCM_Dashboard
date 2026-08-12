import openpyxl

print("=== GESTANTES EXCEL ===")
wb_g = openpyxl.load_workbook('INDICADORES HIS - GESTANTES v1.0 - Junio 2026.xlsx', read_only=True)
print("Gestantes sheets:", wb_g.sheetnames)

print("\n=== NIÑOS EXCEL ===")
wb_n = openpyxl.load_workbook('INDICADORES HIS - NIÑOS v2.1 - Junio 2026.xlsx', read_only=True)
print("Niños sheets:", wb_n.sheetnames)
