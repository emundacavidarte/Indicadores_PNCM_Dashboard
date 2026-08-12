import openpyxl

print("=== VERIFYING GESTANTES TOTALS FOR 202606 ===")
gest_path = "INDICADORES HIS - GESTANTES v1.0 - Junio 2026.xlsx"
wb_gest = openpyxl.load_workbook(gest_path, read_only=True)
ws_g = wb_gest['Tabla']

g_rows = list(ws_g.iter_rows(values_only=True))
header_g = g_rows[0]
print("Gestantes columns:", header_g)

# Map col index
col_g = {name: idx for idx, name in enumerate(header_g)}

g_tot = 0
g_den_anemia, g_num_anemia = 0, 0
g_den_pqt, g_num_pqt = 0, 0
g_den_apn, g_num_apn = 0, 0
g_den_sfaf, g_num_sfaf = 0, 0
g_den_aux, g_num_aux = 0, 0
g_den_parto, g_num_parto = 0, 0

for r in g_rows[1:]:
    p = str(r[col_g['periodo']])
    if p == '202606':
        g_tot += (r[col_g['total_usuarios']] or 0)
        g_num_anemia += (r[col_g['num_anemia']] or 0)
        g_den_anemia += (r[col_g['den_anemia']] or 0)
        g_num_pqt += (r[col_g['num_pqt']] or 0)
        g_den_pqt += (r[col_g['den_pqt']] or 0)
        g_num_apn += (r[col_g['num_apn']] or 0)
        g_den_apn += (r[col_g['den_apn']] or 0)
        g_num_sfaf += (r[col_g['num_sfaf']] or 0)
        g_den_sfaf += (r[col_g['den_sfaf']] or 0)
        g_num_aux += (r[col_g['num_aux']] or 0)
        g_den_aux += (r[col_g['den_aux']] or 0)
        g_num_parto += (r[col_g['num_parto_ins']] or 0)
        g_den_parto += (r[col_g['den_parto_ins']] or 0)

print(f"Gestantes Total 202606: {g_tot}")
print(f"Anemia: num={g_num_anemia}, den={g_den_anemia}, pct={round(g_num_anemia/g_den_anemia*100, 2) if g_den_anemia else 0}%")
print(f"Pqt: num={g_num_pqt}, den={g_den_pqt}, pct={round(g_num_pqt/g_den_pqt*100, 2) if g_den_pqt else 0}%")
print(f"APN: num={g_num_apn}, den={g_den_apn}, pct={round(g_num_apn/g_den_apn*100, 2) if g_den_apn else 0}%")
print(f"SFAF: num={g_num_sfaf}, den={g_den_sfaf}, pct={round(g_num_sfaf/g_den_sfaf*100, 2) if g_den_sfaf else 0}%")
print(f"Auxiliares: num={g_num_aux}, den={g_num_aux}, pct={round(g_num_aux/g_den_aux*100, 2) if g_den_aux else 0}%")
print(f"Parto: num={g_num_parto}, den={g_den_parto}, pct={round(g_num_parto/g_den_parto*100, 2) if g_den_parto else 0}%")


print("\n=== VERIFYING NIÑOS TOTALS FOR 202606 ===")
ninos_path = "INDICADORES HIS - NIÑOS v2.1 - Junio 2026.xlsx"
wb_ninos = openpyxl.load_workbook(ninos_path, read_only=True)
ws_n = wb_ninos['Tabla']

n_rows = list(ws_n.iter_rows(values_only=True))
header_n = n_rows[0]
print("Niños columns:", header_n)

col_n = {name: idx for idx, name in enumerate(header_n)}

n_tot = 0
n_den_bpn, n_num_bpn = 0, 0
n_den_npr, n_num_npr = 0, 0
n_den_hb, n_num_hb = 0, 0
n_den_anemia, n_num_anemia = 0, 0
n_den_cred, n_num_cred = 0, 0
n_den_vrn, n_num_vrn = 0, 0
n_den_hierro, n_num_hierro = 0, 0
n_den_pqt, n_num_pqt = 0, 0
n_den_anemia_fe, n_num_anemia_fe = 0, 0

for r in n_rows[1:]:
    p = str(r[col_n['periodo']])
    if p == '202606':
        n_tot += (r[col_n['total_usuarios']] or 0)
        n_num_bpn += (r[col_n['num_bpn']] or 0)
        n_den_bpn += (r[col_n['den_bpn']] or 0)
        n_num_npr += (r[col_n['num_npr']] or 0)
        n_den_npr += (r[col_n['den_npr']] or 0)
        n_num_hb += (r[col_n['num_hb']] or 0)
        n_den_hb += (r[col_n['den_hb']] or 0)
        n_num_anemia += (r[col_n['num_anemia']] or 0)
        n_den_anemia += (r[col_n['den_anemia']] or 0)
        n_num_cred += (r[col_n['num_cred']] or 0)
        n_den_cred += (r[col_n['den_cred']] or 0)
        n_num_vrn += (r[col_n['num_vrn']] or 0)
        n_den_vrn += (r[col_n['den_vrn']] or 0)
        n_num_hierro += (r[col_n['num_hierro']] or 0)
        n_den_hierro += (r[col_n['den_hierro']] or 0)
        n_num_pqt += (r[col_n['num_pqt']] or 0)
        n_den_pqt += (r[col_n['den_pqt']] or 0)
        n_num_anemia_fe += (r[col_n['num_anemia_fe']] or 0)
        n_den_anemia_fe += (r[col_n['den_anemia_fe']] or 0)

print(f"Niños Total 202606: {n_tot}")
print(f"DNI 30d (BPN): num={n_num_bpn}, den={n_den_bpn}, pct={round(n_num_bpn/n_den_bpn*100, 2) if n_den_bpn else 0}%")
print(f"Recuperación Anemia (NPR): num={n_num_npr}, den={n_den_npr}, pct={round(n_num_npr/n_den_npr*100, 2) if n_den_npr else 0}%")
print(f"Dosaje Hb: num={n_num_hb}, den={n_den_hb}, pct={round(n_num_hb/n_den_hb*100, 2) if n_den_hb else 0}%")
print(f"Frecuencia Anemia: num={n_num_anemia}, den={n_den_anemia}, pct={round(n_num_anemia/n_den_anemia*100, 2) if n_den_anemia else 0}%")
print(f"Control CRED: num={n_num_cred}, den={n_den_cred}, pct={round(n_num_cred/n_den_cred*100, 2) if n_den_cred else 0}%")
print(f"Vacunas VRN: num={n_num_vrn}, den={n_den_vrn}, pct={round(n_num_vrn/n_den_vrn*100, 2) if n_den_vrn else 0}%")
print(f"Suplementación Hierro: num={n_num_hierro}, den={n_den_hierro}, pct={round(n_num_hierro/n_den_hierro*100, 2) if n_den_hierro else 0}%")
print(f"Paquete Integrado: num={n_num_pqt}, den={n_den_pqt}, pct={round(n_num_pqt/n_den_pqt*100, 2) if n_den_pqt else 0}%")
print(f"Tratamiento Hierro: num={n_num_anemia_fe}, den={n_den_anemia_fe}, pct={round(n_num_anemia_fe/n_den_anemia_fe*100, 2) if n_den_anemia_fe else 0}%")
