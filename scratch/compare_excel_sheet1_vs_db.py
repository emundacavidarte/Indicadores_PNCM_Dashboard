import sqlite3
import openpyxl

excel_path = "INDICADORES HIS - NIÑOS v2.1 - Junio 2026.xlsx"
db_path = "dashboard_data.db"

# 1. Read Sheet 1 INDICADORES from Excel
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
ws = wb["INDICADORES"]

# Map period string like 'Ene-23' -> '202301'
month_map = {
    'Ene': '01', 'Feb': '02', 'Mar': '03', 'Abr': '04', 'May': '05', 'Jun': '06',
    'Jul': '07', 'Ago': '08', 'Set': '09', 'Oct': '10', 'Nov': '11', 'Dic': '12'
}

excel_summary = {}

for r in range(12, 54):
    p_str = ws.cell(r, 4).value # Col 4 is PERIODO (e.g. Ene-23)
    if not p_str: continue
    m, y = p_str.split('-')
    p_code = f"20{y}{month_map[m]}"
    
    total_u = ws.cell(r, 5).value
    num_dos = ws.cell(r, 6).value
    den_dos = ws.cell(r, 7).value
    num_an = ws.cell(r, 10).value
    den_an = ws.cell(r, 11).value
    num_cr = ws.cell(r, 14).value
    den_cr = ws.cell(r, 15).value
    num_vr = ws.cell(r, 18).value
    den_vr = ws.cell(r, 19).value
    num_fe = ws.cell(r, 22).value
    den_fe = ws.cell(r, 23).value
    num_vc = ws.cell(r, 26).value
    den_vc = ws.cell(r, 27).value
    num_an_fe = ws.cell(r, 30).value
    den_an_fe = ws.cell(r, 31).value
    
    excel_summary[p_code] = {
        'total_u': int(total_u or 0),
        'num_dos': int(num_dos or 0), 'den_dos': int(den_dos or 0),
        'num_an': int(num_an or 0), 'den_an': int(den_an or 0),
        'num_cr': int(num_cr or 0), 'den_cr': int(den_cr or 0),
        'num_vr': int(num_vr or 0), 'den_vr': int(den_vr or 0),
        'num_fe': int(num_fe or 0), 'den_fe': int(den_fe or 0),
        'num_vc': int(num_vc or 0), 'den_vc': int(den_vc or 0),
        'num_an_fe': int(num_an_fe or 0), 'den_an_fe': int(den_an_fe or 0),
    }

wb.close()

# 2. Read SQLite ninos totals
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

cursor.execute("""
    SELECT periodo,
           SUM(total_usuarios),
           SUM(num_hb), SUM(den_hb),
           SUM(num_anemia), SUM(den_anemia),
           SUM(num_cred), SUM(den_cred),
           SUM(num_vrn), SUM(den_vrn),
           SUM(num_hierro), SUM(den_hierro),
           SUM(num_vac_completa), SUM(den_vac_completa),
           SUM(num_anemia_fe), SUM(den_anemia_fe)
    FROM ninos
    GROUP BY periodo
    ORDER BY periodo
""")

db_summary = {}
for r in cursor.fetchall():
    p_code = str(r[0])
    db_summary[p_code] = {
        'total_u': int(r[1] or 0),
        'num_dos': int(r[2] or 0), 'den_dos': int(r[3] or 0),
        'num_an': int(r[4] or 0), 'den_an': int(r[5] or 0),
        'num_cr': int(r[6] or 0), 'den_cr': int(r[7] or 0),
        'num_vr': int(r[8] or 0), 'den_vr': int(r[9] or 0),
        'num_fe': int(r[10] or 0), 'den_fe': int(r[11] or 0),
        'num_vc': int(r[12] or 0), 'den_vc': int(r[13] or 0),
        'num_an_fe': int(r[14] or 0), 'den_an_fe': int(r[15] or 0),
    }

conn.close()

# 3. Compare Excel Sheet 1 vs SQLite DB
all_periods = sorted(list(set(list(excel_summary.keys()) + list(db_summary.keys()))))

print(f"Comparing Sheet 1 'INDICADORES' ({len(excel_summary)} periods) vs SQLite 'ninos' ({len(db_summary)} periods)...\n")

diffs = []
for p in all_periods:
    ex = excel_summary.get(p)
    db = db_summary.get(p)
    if not ex:
        diffs.append(f"Period {p} in DB but missing from Excel Sheet 1")
        continue
    if not db:
        diffs.append(f"Period {p} in Excel Sheet 1 but missing from DB")
        continue
    
    for k in ex.keys():
        if ex[k] != db[k]:
            diffs.append(f"Period {p} | Field {k:<10} | Excel Sheet 1: {ex[k]:10,d} | DB: {db[k]:10,d} | Diff: {db[k]-ex[k]:+10,d}")

if diffs:
    print(f"⚠️ FOUND {len(diffs)} DISCREPANCIES BETWEEN EXCEL SHEET 1 AND SQLITE DB:")
    for d in diffs:
        print("  ", d)
else:
    print("[OK] PERFECT MATCH! All 42 periods and all indicator totals match 100% between Excel Sheet 1 'INDICADORES' and SQLite database 'ninos'!")
