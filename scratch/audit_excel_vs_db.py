import openpyxl
import sqlite3

def check_excel_totals():
    conn = sqlite3.connect('dashboard_data.db')
    cursor = conn.cursor()

    print("=== GESTANTES EXCEL vs DB ===")
    wb_g = openpyxl.load_workbook('INDICADORES HIS - GESTANTES v1.0 - Junio 2026.xlsx', data_only=True, read_only=True)
    ws_g_tabla = wb_g['Tabla']
    
    g_headers = None
    g_count = 0
    g_sum_padron = 0
    g_sum_num_anemia = 0
    g_sum_den_anemia = 0
    
    for r in ws_g_tabla.iter_rows(values_only=True):
        if not g_headers:
            g_headers = [str(c).upper() if c is not None else '' for c in r]
            print("Gestantes Tabla Headers:", g_headers)
            continue
        g_count += 1
        # Sum columns for 202606 if PERIODO is column 0
        periodo = str(r[0]) if r[0] is not None else ''
        if periodo == '202606':
            # find index of TOTAL_USUARIOS/PADRON, NUM_ANEMIA, DEN_ANEMIA
            # Let's print row 2 first
            if g_count == 1:
                print("Gestantes Sample Row:", r[:15])

    print(f"Total Gestantes Tabla rows: {g_count}")

    print("\n=== NIÑOS EXCEL vs DB ===")
    wb_n = openpyxl.load_workbook('INDICADORES HIS - NIÑOS v2.1 - Junio 2026.xlsx', data_only=True, read_only=True)
    ws_n_tabla = wb_n['Tabla']
    
    n_headers = None
    n_count = 0
    for r in ws_n_tabla.iter_rows(values_only=True):
        if not n_headers:
            n_headers = [str(c).upper() if c is not None else '' for c in r]
            print("Niños Tabla Headers:", n_headers[:15])
            continue
        n_count += 1
        if n_count == 1:
            print("Niños Sample Row:", r[:15])

    print(f"Total Niños Tabla rows: {n_count}")
    conn.close()

check_excel_totals()
