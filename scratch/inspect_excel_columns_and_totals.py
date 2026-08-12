import openpyxl

print("=== GESTANTES EXCEL - TABLA COLUMNS ===")
gest_path = "INDICADORES HIS - GESTANTES v1.0 - Junio 2026.xlsx"
wb_gest = openpyxl.load_workbook(gest_path, read_only=True)
ws_g_tabla = wb_gest['Tabla']
header_g = [cell for cell in next(ws_g_tabla.iter_rows(max_row=1, values_only=True))]
print("Gestantes Tabla columns:", header_g)

print("\n=== GESTANTES EXCEL - INDICADORES SHEET ===")
ws_g_ind = wb_gest['INDICADORES']
for r in list(ws_g_ind.iter_rows(values_only=True)):
    non_empty = [c for c in r if c is not None]
    if non_empty:
        print("  ", non_empty[:8])

print("\n=== NIÑOS EXCEL - TABLA COLUMNS ===")
ninos_path = "INDICADORES HIS - NIÑOS v2.1 - Junio 2026.xlsx"
wb_ninos = openpyxl.load_workbook(ninos_path, read_only=True)
ws_n_tabla = wb_ninos['Tabla']
header_n = [cell for cell in next(ws_n_tabla.iter_rows(max_row=1, values_only=True))]
print("Niños Tabla columns:", header_n)

print("\n=== NIÑOS EXCEL - INDICADORES SHEET ===")
ws_n_ind = wb_ninos['INDICADORES']
for r in list(ws_n_ind.iter_rows(values_only=True)):
    non_empty = [c for c in r if c is not None]
    if non_empty:
        print("  ", non_empty[:8])
