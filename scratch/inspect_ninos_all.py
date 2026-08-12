import sqlite3
import openpyxl
import os

excel_path = "INDICADORES HIS - NIÑOS v2.1 - Junio 2026.xlsx"
db_path = "dashboard_data.db"

print("--- EXCEL FILE INSPECTION ---")
print(f"File: {excel_path}")
print(f"Size: {os.path.getsize(excel_path) / (1024*1024):.2f} MB")

wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=False)
print(f"Sheet names ({len(wb.sheetnames)}): {wb.sheetnames}")

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\nSheet: '{sheetname}' | max_row={ws.max_row}, max_column={ws.max_column}")
    row_count = 0
    for row in ws.iter_rows(values_only=True):
        row_count += 1
        if row_count <= 5:
            row_str = [str(c)[:40] if c is not None else "" for c in row[:15]]
            print(f"  Row {row_count}: {row_str}")
        if row_count > 5:
            break

wb.close()

print("\n--- SQLITE DATABASE INSPECTION ---")
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
tables = [r[0] for r in cursor.fetchall()]
print(f"Tables in DB: {tables}")

for tbl in ['ninos', 'ninos_summary', 'ninos_geo_summary']:
    if tbl in tables:
        cursor.execute(f"PRAGMA table_info({tbl})")
        cols = cursor.fetchall()
        print(f"\nTable: {tbl} ({len(cols)} columns)")
        col_names = [c[1] for c in cols]
        print("  Columns:", col_names)
        
        cursor.execute(f"SELECT COUNT(*) FROM {tbl}")
        cnt = cursor.fetchone()[0]
        print(f"  Total rows: {cnt:,}")

conn.close()
