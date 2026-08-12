import openpyxl

print("=== GESTANTES EXCEL - INDICADORES SHEET ===")
wb_g = openpyxl.load_workbook('INDICADORES HIS - GESTANTES v1.0 - Junio 2026.xlsx', data_only=True, read_only=True)
ws_g = wb_g['INDICADORES']

headers_g = []
for r in ws_g.iter_rows(min_row=1, max_row=5, values_only=True):
    print("Row:", [str(cell)[:30] for cell in r if cell is not None][:15])

print("\n=== NIÑOS EXCEL - INDICADORES SHEET ===")
wb_n = openpyxl.load_workbook('INDICADORES HIS - NIÑOS v2.1 - Junio 2026.xlsx', data_only=True, read_only=True)
ws_n = wb_n['INDICADORES']

for r in ws_n.iter_rows(min_row=1, max_row=5, values_only=True):
    print("Row:", [str(cell)[:30] for cell in r if cell is not None][:15])
