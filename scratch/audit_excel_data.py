import openpyxl

print("=== 1. INSPECTING GESTANTES EXCEL ===")
gest_path = "INDICADORES HIS - GESTANTES v1.0 - Junio 2026.xlsx"
wb_gest = openpyxl.load_workbook(gest_path, read_only=True)
print("Gestantes sheets:", wb_gest.sheetnames)

for sname in wb_gest.sheetnames:
    ws = wb_gest[sname]
    rows = list(ws.iter_rows(max_row=5, values_only=True))
    print(f"\nSheet '{sname}' max_row={ws.max_row}, max_column={ws.max_column}:")
    for r in rows[:3]:
        print("  ", r[:10])

print("\n=== 2. INSPECTING NIÑOS EXCEL ===")
ninos_path = "INDICADORES HIS - NIÑOS v2.1 - Junio 2026.xlsx"
wb_ninos = openpyxl.load_workbook(ninos_path, read_only=True)
print("Niños sheets:", wb_ninos.sheetnames)

for sname in wb_ninos.sheetnames:
    ws = wb_ninos[sname]
    rows = list(ws.iter_rows(max_row=5, values_only=True))
    print(f"\nSheet '{sname}' max_row={ws.max_row}, max_column={ws.max_column}:")
    for r in rows[:3]:
        print("  ", r[:10])
